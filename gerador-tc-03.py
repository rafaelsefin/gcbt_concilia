import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.cell_range import CellRange
import os
import shutil
from zipfile import ZipFile
from datetime import datetime

# Diretórios de trabalho
OUTPUT_DIR = "output_files"
ZIP_PATH = "planilhas_geradas.zip"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Mapeamento de meses
MESES_PT = {
    "January": "JANEIRO", "February": "FEVEREIRO", "March": "MARÇO",
    "April": "ABRIL", "May": "MAIO", "June": "JUNHO", "July": "JULHO",
    "August": "AGOSTO", "September": "SETEMBRO", "October": "OUTUBRO",
    "November": "NOVEMBRO", "December": "DEZEMBRO"
}

def formatar_conta(valor):
    valor = str(valor).strip()
    if "-" in valor or not valor.isdigit():
        return valor
    return f"{valor[:-1]}-{valor[-1]}"

def gerar_planilhas(dados, modelo):
    df = pd.read_excel(dados, header=1)
    df.columns = df.columns.str.strip()
    df = df.rename(columns={
        "CONTA": "CONTA_NOME",
        "DISPONÍVEL EM CONTA CORRENTE": "CONTA_CORRENTE",
        "APLICAÇÃO FINANCEIRA": "APLICACAO",
        "REGISTRADO/SIGEF": "SIGEF"
    })

    for idx, row in df.iterrows():
        try:
            conta = formatar_conta(row["CONTA_NOME"])
            val_c12 = float(row["CONTA_CORRENTE"])
            val_c13 = float(row["APLICACAO"])
            val_c14 = float(row["SIGEF"])

            data = pd.to_datetime(row["DATA"], errors="coerce")
            if pd.isna(data):
                mes_pt, ano, data_fmt, local_data = "", "", "", ""
            else:
                mes_pt = MESES_PT.get(data.strftime("%B"), data.strftime("%B").upper())
                ano = data.strftime("%Y")
                data_fmt = data.strftime("%d/%m/%Y")
                local_data = f"Porto Velho, {data_fmt}"

            wb = load_workbook(modelo)
            ws = wb["Dados"]

            targets = {
                "C8": conta,
                "C9": mes_pt,
                "C10": ano,
                "C12": val_c12,
                "C13": val_c13,
                "C14": val_c14,
                "C16": data_fmt,
                "C21": local_data
            }

            to_unmerge = []
            for merged in ws.merged_cells.ranges:
                cr = CellRange(str(merged))
                for cell in targets:
                    r = int(cell[1:])
                    c = column_index_from_string(cell[0])
                    if cr.min_row <= r <= cr.max_row and cr.min_col <= c <= cr.max_col:
                        to_unmerge.append(str(merged))
            for rng in set(to_unmerge):
                ws.unmerge_cells(rng)

            for cell, val in targets.items():
                ws[cell] = val

            wb.save(f"{OUTPUT_DIR}/{conta}.xlsx")

        except Exception as e:
            st.error(f"Erro na linha {idx + 2}: {e}")

    # Zip final
    with ZipFile(ZIP_PATH, 'w') as zipf:
        for file in os.listdir(OUTPUT_DIR):
            zipf.write(os.path.join(OUTPUT_DIR, file), arcname=file)

# Streamlit UI
st.set_page_config(page_title="Gerador de Planilhas", layout="centered")
st.title("📊 Gerador de Planilhas de Conciliação")

col1, col2 = st.columns(2)
with col1:
    dados_file = st.file_uploader("📥 Envie a planilha de dados", type="xlsx", key="dados")
with col2:
    modelo_file = st.file_uploader("📥 Envie o modelo", type="xlsx", key="modelo")

if st.button("🚀 Gerar Planilhas"):
    if dados_file and modelo_file:
        with open("dados_temp.xlsx", "wb") as f:
            f.write(dados_file.read())
        with open("modelo_temp.xlsx", "wb") as f:
            f.write(modelo_file.read())

        gerar_planilhas("dados_temp.xlsx", "modelo_temp.xlsx")
        st.success("✅ Planilhas geradas com sucesso!")

        with open(ZIP_PATH, "rb") as f:
            st.download_button("⬇️ Baixar Arquivos Zipados", f, file_name="planilhas_concilia.zip")
    else:
        st.warning("Por favor envie os dois arquivos.")

if st.button("🧹 Limpar Base"):
    shutil.rmtree(OUTPUT_DIR, ignore_errors=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if os.path.exists(ZIP_PATH):
        os.remove(ZIP_PATH)
    st.info("Base limpa com sucesso!")
